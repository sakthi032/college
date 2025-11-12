import io
import json
import os
from datetime import datetime
import pandas as pd
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Font, Border, Side, Alignment
from django.core.files.base import ContentFile
from .forms import RegisterForm
from django.contrib.auth.models import User
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from .models import Department, StaffProfile, StudentGrade, ExcelFile, InternalExcelFile, CollegeExam, CollegeExamExcel, SendMergeFile, MergeFile


def front_page(request):
    return render(request,'staff/frontpage.html')


def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user=form.save()
            messages.success(request, f"Account created successfully by {user.username}!")
            login(request, user)
            return redirect("add_dept")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = RegisterForm()
    return render(request, "staff/register.html", {"form": form})

def login_view(request):
    if request.method == "POST":
        identifier = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        user = None
        if "@" in identifier:
            try:
                user_obj = User.objects.get(email=identifier)
                username = user_obj.username
            except User.DoesNotExist:
                username = None
        else:
            username = identifier
        if username:
            user = authenticate(request, username=username, password=password)

        if user is not None:
            if not user.email or user.email.strip() == "":
                messages.error(request, "This is admin page. Staff are not allowed to log in here.")
                return redirect("login")
            login(request, user)
            messages.success(request, f"Welcome Back {user.username}")
            return redirect("add_dept")
        else:
            messages.error(request, "Invalid username/email or password")
    return render(request, "staff/login.html")


@login_required
def add_department(request):
    if request.method == "POST":
        dept_name = request.POST.get("departments").strip()

        # Check if this admin already has this department
        if Department.objects.filter(user=request.user, name=dept_name).exists():
            messages.error(request, "Department already exists for your account!")
            return redirect("add_dept")

        # Create department
        dept = Department.objects.create(name=dept_name, user=request.user)

        # Generate unique staff username (global uniqueness)
        base_username = dept_name.lower().replace(" ", "_")
        username = f"{request.user.username}_{base_username}"

        if User.objects.filter(username=username).exists():
            messages.warning(request, f"Default staff user '{username}' already exists.")
        else:
            user = User.objects.create_user(
                username=username,
                password="password"  # default password
            )
            StaffProfile.objects.create(user=user, department=dept)
            messages.success(request, f"Department '{dept_name}' created with staff username '{username}'.")

        return redirect("add_dept")

    # For GET request
    departments = Department.objects.filter(user=request.user).order_by("-created_at")
    return render(request, "staff/add_department.html", {"data": departments})

def staff_login(request):
    if request.method == "POST":
        user_name = request.POST.get("username", "").strip()   # admin username
        dept_name = request.POST.get("department", "").strip()
        password = request.POST.get("password", "").strip()

        try:
            # get the admin user object
            admin_user = User.objects.get(username=user_name)
        except User.DoesNotExist:
            messages.error(request, "Invalid admin username!")
            return redirect("staff_login")

        try:
            # get department belonging to that admin
            department = Department.objects.get(user=admin_user, name=dept_name)
        except Department.DoesNotExist:
            messages.error(request, "Invalid department for this admin!")
            return redirect("staff_login")

        # build staff username (like staffname_department)
        staff_username = f"{user_name}_{department.name.strip().lower().replace(' ', '_')}"

        user = authenticate(request, username=staff_username, password=password)

        if user is not None:
            if hasattr(user, "staffprofile") and user.staffprofile.department == department:
                login(request, user)
                return redirect("department")  # staff dashboard
            else:
                messages.error(request, "You are not assigned to this department!")
        else:
            messages.error(request, "Invalid username or password!")

    return render(request, "staff/staff_login.html")



@login_required
def department(request):
    if not request.user.is_authenticated:
        return redirect("staff_login")
    staff_profile = request.user.staffprofile
    department_name = staff_profile.department.name
    category = 'theory'
    if request.method == 'POST':
        search = request.POST.get('search', "").strip()
        category = request.POST.get('department', "theory")
    else:
        search = ""
    # Select queryset based on category
    if category == 'theory':
        qs = ExcelFile.objects.filter(user=request.user).order_by('-id')
    elif category == 'internal':
        qs = InternalExcelFile.objects.filter(user=request.user).order_by('-id')
    elif category == 'college_exam':
        qs = CollegeExamExcel.objects.filter(user=request.user).order_by('-id')
    else:
        qs = []
    if search:
        qs = qs.filter(course_code__icontains=search)
    data = qs
    return render(request,'staff/department.html',context={"category":category,"dept": department_name,"data": data,"selected_category": category.upper().replace('_'," "),'value':search})


@login_required
def forget_password(request):
    return render(request,'staff/forget_password.html')

@login_required
def download_excel_file(request, model_name, file_id):
    model_map = {
        'MergeFile': MergeFile,
        'SendMergeFile': SendMergeFile,
        'CollegeExamExcel': CollegeExamExcel,
        'InternalExcelFile': InternalExcelFile,
        'ExcelFile': ExcelFile,
    }

    model = model_map.get(model_name)
    if not model:
        raise Http404("Invalid file type")

    try:
        excel_file = model.objects.get(id=file_id, user=request.user)
    except model.DoesNotExist:
        raise Http404("File not found")

    response = HttpResponse(
        excel_file.file_data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{excel_file.file_name}"'
    return response

@login_required
def theory(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            header_data = data.get('headerData', {})
            student_data = data.get('students', [])
            ese_data = data.get('eseData', {})
            co_survey_data = data.get('coSurveyData', [])
            overall_data = data.get('overallData', {})
            overall_attainment = data.get('overallAttainment', [])
            course_name = header_data.get('courseName', '').strip().upper()
            course_cod = header_data.get('courseCode', '').strip().upper()
            semester = header_data.get('semester', '').strip()

            for student in student_data:
                reg_no = student.get('regNo', '')
                student_name = student.get('name', '').upper()
                ese = student.get('ese', 0)
                cia = student.get('cia', 0)
                total = student.get('total', 0)

                sts = ""
                log = 0
                if (ese>= 30 and cia >= 10) or (cia<10 and ese>=40):
                    sts = "Pass"
                else:
                    sts = "Fail"
                if sts == 'Pass':
                    log = 0
                else:
                    log = 1

                try:
                    obj, created = StudentGrade.objects.update_or_create(
                        user=request.user,
                        reg_no=reg_no,
                        course_code=course_cod,
                        semester=semester,
                        defaults={
                            "student_name": student_name,
                            "course_name": course_name,
                            "ese": ese,
                            "cia": cia,
                            "total": total,
                            "status":sts,
                            "semester_log":log,
                        }
                    )

                    #save to Excel file
                    wb = Workbook()
                    sheet = wb.active
                    course_code = header_data.get('courseCode', 'Sheet1')
                    sheet.title = course_code

                    times_new_roman_font = Font(name='Times New Roman', size=12)

                    # Apply the font to all cells in the sheet
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.font = times_new_roman_font

                    column_widths = {
                        'A': 3.86, 'B': 15.86, 'C': 42.29, 'D': 10.29, 'E': 8.14,
                        'F': 8.14, 'G': 8.43, 'H': 22.71, 'I': 13, 'J': 8.43,
                        'K': 8.43, 'L': 8.43, 'M': 8.43, 'N': 14.29
                    }
                    for col, width in column_widths.items():
                        sheet.column_dimensions[col].width = width
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Define center alignment
                    center_alignment = Alignment(horizontal='center', vertical='center')

                    bold_font = Font(bold=True)

                    sheet.merge_cells('A2:C2')
                    sheet.merge_cells('A3:C3')
                    sheet.merge_cells('A4:C4')
                    sheet.merge_cells('D2:E2')
                    sheet.merge_cells('D3:E3')
                    sheet.merge_cells('D4:E4')

                    sheet['A2'] = f'Programme: {header_data["programme"].upper()}'
                    sheet['A2'].font = bold_font
                    # sheet['A2'].alignment = center_alignment

                    sheet['A3'] = f'Course Name: {header_data["courseName"].upper()}'
                    sheet['A3'].font = bold_font
                    # sheet['A3'].alignment = center_alignment

                    sheet['A4'] = f'Course Code: {header_data["courseCode"].upper()}'
                    sheet['A4'].font = bold_font
                    # sheet['A4'].alignment = center_alignment

                    sheet['D2'] = "Academic Year :"
                    sheet['D2'].font = bold_font
                    # sheet['D2'].alignment = center_alignment

                    sheet['D3'] = 'Semester :'
                    sheet['D3'].font = bold_font
                    # sheet['D3'].alignment = center_alignment

                    sheet['F2'] = header_data.get('academicYear', '')
                    sheet['F2'].alignment = center_alignment
                    sheet['F3'] = header_data.get('semester', '')
                    sheet['F3'].alignment = center_alignment

                    merged_cells = ['A2:C2', 'A3:C3', 'A4:C4', 'D2:E2', 'D3:E3', 'D4:E4']
                    for merged_cell in merged_cells:
                        for row in sheet[merged_cell]:
                            for cell in row:
                                cell.border = thin_border
                    for row in sheet['A2:F4']:
                        for cell in row:
                            cell.border = thin_border

                    # Add an empty row for spacing
                    sheet.append([])
                    # Add Student Data
                    sheet.append(["SNO", "Reg No", "Name of The Students", "ESE", "CIA", "Total"])

                    # Apply bold font to header row
                    sheet['A6'].font = bold_font
                    sheet['B6'].font = bold_font
                    sheet['C6'].font = bold_font
                    sheet['D6'].font = bold_font
                    sheet['E6'].font = bold_font
                    sheet['F6'].font = bold_font

                    # Apply center alignment for all headers except student name column
                    for col in ['A', 'B', 'D', 'E', 'F']:
                        cell = sheet[f'{col}6']
                        cell.alignment = Alignment(horizontal='center')

                    # Add student data
                    for idx, student in enumerate(student_data, start=1):
                        sheet.append([
                            idx,
                            student.get('regNo', ''),
                            student.get('name', '').upper(),
                            student.get('ese', 0),
                            student.get('cia', 0),
                            student.get('total', 0)
                        ])

                    end_row = len(student_data) + 6
                    for row in sheet[f'A6:F{end_row}']:
                        for cell in row:
                            cell.border = thin_border

                    # Apply alignment for student name column (starting from C7 onwards)
                    for row in sheet[f'C7:C{end_row}']:
                        for cell in row:
                            cell.alignment = Alignment(horizontal='left')

                    # Apply center alignment to other columns
                    for col in ['A', 'B', 'D', 'E', 'F']:
                        for row in sheet[f'{col}7:{col}{end_row}']:
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center')

                    sheet.append([])

                    # Add ESE Grade Distribution (starting from row 6)
                    sheet.merge_cells('H6:I6')
                    sheet['H6'] = 'Direct Tool - University Examination [ESE]'
                    sheet['H6'].font = bold_font
                    sheet['H6'].alignment = center_alignment

                    sheet['H7'] = 'Total No of "O" Grade'
                    sheet['H8'] = 'Total No of "D" Grade'
                    sheet['H9'] = 'Total No of "A+" Grade'
                    sheet['H10'] = 'Total No of "A" Grade'
                    sheet['H11'] = 'Total No of "B" Grade'
                    sheet['H12'] = 'Total No of "C" Grade'
                    sheet['H13'] = 'Total No of "U" Grade'
                    sheet['H14'] = 'Total No of "AAA" Grade'
                    sheet['H17'] = 'Total Students'
                    sheet['H19'] = 'No of Passed'
                    sheet['H20'] = 'Percentage "%"'
                    sheet['H21'] = 'UNIV ATTAINMENT'

                    # Add values for ESE Data
                    sheet['I7'] = ese_data.get('O')
                    sheet['I8'] = ese_data.get('D')
                    sheet['I9'] = ese_data.get('A+')
                    sheet['I10'] = ese_data.get('A')
                    sheet['I11'] = ese_data.get('B')
                    sheet['I12'] = ese_data.get('C')
                    sheet['I13'] = ese_data.get('U')
                    sheet['I14'] = ese_data.get('AAA')
                    sheet['I17'] = ese_data.get('totalStudents')
                    sheet['I19'] = ese_data.get('passedStudents')
                    sheet['I20'] = ese_data.get('percentage')
                    sheet['I21'] = ese_data.get('universityAttainment')

                    # Apply borders to ESE section
                    for row in sheet['H6:I21']:
                        for cell in row:
                            cell.border = thin_border

                    sheet.append([])

                    # Add Overall Data
                    sheet.merge_cells('H23:K23')
                    sheet['H23'] = 'Direct Tool - Internal Examination [CIA]'
                    sheet['H23'].font = bold_font
                    sheet['H23'].alignment = center_alignment

                    sheet['H25'] = 'More than 70%'
                    sheet['I25'] = overall_data.get('cia70', 0)
                    sheet['J25'] = overall_data.get('ciaper', 0)
                    sheet['K25'] = overall_data.get('ciauni', 0)

                    # Apply borders to CIA section
                    for row in sheet['H23:K25']:
                        for cell in row:
                            cell.border = thin_border

                    sheet.append([])

                    # Add CO Survey Data
                    sheet.merge_cells('H28:K29')
                    sheet.merge_cells('L28:L30')
                    sheet.merge_cells('M28:M30')
                    sheet.merge_cells('N28:N30')

                    sheet['H28'] = 'Indirect Tool - Course Outcome Survey'
                    sheet['L28'] = 'Total'
                    sheet['M28'] = 'Percentage %'
                    sheet["N28"] = 'Outcome'
                    sheet['H30'] = 'COs'
                    sheet['I30'] = 'Excellent'
                    sheet['J30'] = 'Good'
                    sheet['K30'] = 'Fair'

                    # Apply bold font and center alignment
                    sheet['H28'].font = bold_font
                    sheet['L28'].font = bold_font
                    sheet['M28'].font = bold_font
                    sheet['N28'].font = bold_font

                    # Fill in the CO Survey Data
                    for row, co in zip(range(31, 36), co_survey_data):
                        sheet[f'H{row}'] = co.get('co', '')
                        sheet[f'I{row}'] = co.get('excellent', 0)
                        sheet[f'J{row}'] = co.get('good', 0)
                        sheet[f'K{row}'] = co.get('fair', 0)
                        sheet[f'L{row}'] = co.get('total', 0)
                        sheet[f'M{row}'] = co.get('percentage', 0)
                        sheet[f'N{row}'] = co.get('outcome', '')

                    # Apply borders to CO Survey section
                    for row in sheet['H28:N36']:
                        for cell in row:
                            cell.border = thin_border

                    sheet.append([])

                    # Apply formatting for the Overall Attainment section
                    sheet['H39'] = 'Overall CO Attainment'
                    sheet['H40'] = 'COs'
                    sheet['I40'] = 'Survey(IDA)'
                    sheet['J40'] = '0.2 * IDA'
                    sheet['K40'] = '0.6 * U A'
                    sheet['L40'] = '0.4 * IA'
                    sheet['M40'] = 'DA'
                    sheet['N40'] = 'Overall Attainment'
                    sheet['H39'].font = bold_font

                    # Apply bold font and center alignment to headers
                    for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N']:
                        sheet[f'{col}40'].font = bold_font
                        sheet[f'{col}40'].alignment = center_alignment

                        # Add Overall Attainment Data (Rows 41 to 46 for CO1 to CO5)
                        for row, attainment in zip(range(41, 46), overall_attainment):
                            sheet[f'H{row}'] = attainment.get('co', '')
                            sheet[f'I{row}'] = attainment.get('surveyIDA', 0)
                            sheet[f'J{row}'] = attainment.get('point1', 0)
                            sheet[f'K{row}'] = attainment.get('point21', 0)
                            sheet[f'L{row}'] = attainment.get('ia', 0)
                            sheet[f'M{row}'] = attainment.get('totalPoint', 0)
                            sheet[f'N{row}'] = attainment.get('max', 0)

                        # Add Summary Data (Once, after the rows for COs)
                        sheet['I47'] = overall_attainment[0].get('allPoint', 0)
                        sheet['I48'] = 'Overall Attainment'
                        sheet['M47'] = overall_attainment[0].get('average', 0)
                        sheet['N47'] = overall_attainment[0].get('allTotal', 0)
                        sheet['N48'] = overall_attainment[0].get('ultraMax', 0)

                        # Apply borders to Overall Attainment section (Rows 40 to 48)
                        for row in sheet['H39:N48']:
                            for cell in row:
                                cell.border = thin_border

                    # Merge footer cells and add labels
                    merge_ranges = ['H50:J50', 'H51:J51', 'H52:J52', 'H53:J53', 'H57:K57']
                    for merge_range in merge_ranges:
                        sheet.merge_cells(merge_range)

                    # Set values and apply the bold font
                    sheet['H50'] = 'ESE - End Semester Examination'
                    sheet['H50'].font = bold_font

                    sheet['H51'] = 'IDA - In Direct Attainment'
                    sheet['H51'].font = bold_font

                    sheet['H52'] = 'DA - Direct Attainment'
                    sheet['H52'].font = bold_font

                    sheet['H53'] = 'IA - Internal Examination Attainment'
                    sheet['H53'].font = bold_font

                    sheet['H57'] = 'Signature of Course I/C'
                    sheet['H57'].font = bold_font

                    sheet['N57'] = 'HOD'
                    sheet['N57'].font = bold_font
                    course_code = header_data.get('courseCode', 'default')
                    filename = f"{course_code}.xlsx"
                    stream = io.BytesIO()
                    wb.save(stream)
                    file_bytes = stream.getvalue()

                    # Step 2: Save or update Excel file in DB
                    excel_file, created = ExcelFile.objects.update_or_create(
                        user=request.user,
                        course_code=course_code,
                        file_name=filename,
                        defaults={"file_data": file_bytes}
                    )
                except Exception as e:
                    print(f"Error saving student {reg_no}: {e}")
            return JsonResponse({"status": "success", "message": "Data saved successfully!"})
        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return render(request,'staff/theory_form.html')


@login_required
def internal(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            wb = Workbook()
            sheet = wb.active
            header_data = data.get('headerData', {})
            student_data = data.get('students', [])
            course_code = header_data.get('courseCode', 'Sheet1')
            sheet.title = course_code
            times_new_roman_font = Font(name='Times New Roman', size=11)

            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = times_new_roman_font

            # Column width settings
            column_widths = {
                'A': 5.5, 'B': 13.6, 'C': 30, 'D': 7.5, 'E': 7.5,
                'F': 9, 'G': 7.5, 'H': 10, 'I': 8.8, 'J': 8.8,
                'K': 8.8, 'L': 8.8, 'M': 8.8, 'N': 14.29
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Define border and styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            bold_font = Font(bold=True)
            align_left_center = Alignment(horizontal="left", vertical="center")
            # Set row heights
            sheet.row_dimensions[3].height = 25
            sheet.row_dimensions[5].height = 45

            # Merge cells for headers
            sheet.merge_cells('A3:C3')
            sheet.merge_cells('D3:G3')
            sheet.merge_cells('H3:L3')
            sheet.merge_cells('M3:N3')

            # Set values for headers
            sheet['A3'] = f'Programme: {header_data["programme"].upper()}'
            sheet['A3'].font = bold_font
            sheet['A3'].alignment = align_left_center  # Vertical center, Left start

            sheet['D3'] = f'Year & Sem: {header_data["academicYear"].upper()} & {header_data["semester"].upper()}'
            sheet['D3'].font = bold_font
            sheet['D3'].alignment = align_left_center

            sheet['H3'] = f'Course Name: {header_data["courseName"].upper()}'
            sheet['H3'].font = bold_font
            sheet['H3'].alignment = align_left_center

            sheet['M3'] = f'Course Code: {header_data["courseCode"].upper()}'
            sheet['M3'].font = bold_font
            sheet['M3'].alignment = align_left_center

            # Append header row
            sheet.append([])
            header_row = 5
            for col in range(1, 15):  # Columns A to N (1 to 14)
                cell = sheet.cell(row=header_row, column=col)
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col in range(1, 15):  # Columns A(1) to N(14)
                sheet.cell(row=3, column=col).border = thin_border

            # Apply border to each cell in range A5:N5
            for col in range(1, 15):  # Columns A(1) to N(14)
                sheet.cell(row=5, column=col).border = thin_border

            # Modify header text to add line breaks before parentheses
            header_values = [
                "S.No", "Reg No", "Name", "No of\ndays present", "No of\nworking days", "Percentage",
                "Marks\n(5)", "Assignment\n(5)", "Seminar\n(5)", "I Internal\n(50)", "II Internal\n(50)",
                "Model Exam\n(75)", "Total\n(10)", "Overall Total\n(25)"
            ]

            for col, value in enumerate(header_values, start=1):
                sheet.cell(row=header_row, column=col, value=value)
            start_row = 6
            for idx, student in enumerate(student_data, start=1):
                row_num=start_row + idx -1
                sheet.append([
                    idx,  # Serial Number
                    student.get('reg_no', ''),
                    student.get('student_name', '').upper(),
                    student.get('days_present', 0),
                    student.get('working_days', 0),
                    student.get('percentage', 0),
                    student.get('marks', 0),
                    student.get('assignment', 0),
                    student.get('seminar', 0),
                    student.get('internal1', 0),
                    student.get('internal2', 0),
                    student.get('model_exam', 0),
                    student.get('total', 0),
                    student.get('overall_total', 0),
                ])

                for col in range(1,15):
                    cell=sheet.cell(row=row_num,column=col)
                    cell.border=thin_border

            last_student_row = start_row + len(student_data) - 1
            empty_row_1 = last_student_row + 1
            empty_row_2 = empty_row_1 + 1
            final_row = empty_row_2 + 1
            sheet.append([])
            sheet.append([])

            sheet.append(["Faculty", "", "", "", "", "HOD", "", "", "", "", "", "Principal", "",""])

            # Merge cells for signatures
            sheet.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=2)  # Merge A-B
            sheet.merge_cells(start_row=final_row, start_column=6, end_row=final_row, end_column=7)  # Merge F-G
            sheet.merge_cells(start_row=final_row, start_column=12, end_row=final_row, end_column=14)  # Merge L-N

            for col in [1, 6, 12]:  # Faculty (A-B), HOD (F-G), Principal (L-N)
                for c in range(col, col + 2 if col != 12 else col + 3):  # A-B, F-G, L-N
                    cell = sheet.cell(row=final_row, column=c)
                    cell.border = thin_border  # Apply thin border

            for row in range(6, last_student_row + 1):
                for col in range(1, 15):
                    cell = sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            course_code = header_data.get('courseCode', 'default')
            filename = f"{course_code}.xlsx"

            stream = io.BytesIO()
            wb.save(stream)
            excel_bytes = stream.getvalue()

            obj, created = InternalExcelFile.objects.update_or_create(
            user=request.user,
            course_code=course_code,
            defaults={
                "file_name": filename,
                "file_data": excel_bytes,
            }
        )
            return JsonResponse({"status": "success", "message": "Data saved successfully!"})
        except Exception as e:
            print("Error:", str(e))
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return render(request,'staff/internal_form.html')


@login_required
def college_exam(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            header = data.get('headerData', {})
            students = data.get('studentsData', [])
            for student in students:
                sts = ""
                log=0
                if int(student.get("total"))==50 and  int(student.get("marks"))>18 :
                    sts="Pass"
                elif int(student.get("total"))==75 and  int(student.get("marks"))>30 :
                    sts="Pass"
                else:
                    sts="Fail"
                if sts=='Pass':
                    log=0
                else:
                    log=1

                CollegeExam.objects.update_or_create(
                    user=request.user,
                    exam_name=header.get("examName", "").upper(),
                    course_code=header.get("courseCode", "").upper(),
                    reg_no=student.get("regNo", ""),
                    defaults={
                        'programme': header.get("programme", "").upper(),
                        'course_name': header.get("courseName", "").upper(),
                        'academic_year': header.get("academicYear", "").upper(),
                        'semester': header.get("semester", "").upper(),
                        'student_name': student.get("studentName", "").upper(),
                        'marks': int(student.get("marks", 0)),
                        'total': int(student.get("total", 0)),
                        'percentage': float(student.get("percentage", 0)),
                        'status':sts,
                        'college_log':log,
                    }
                )

            # Create Excel workbook
            wb = Workbook()
            sheet = wb.active
            course_code = header.get('courseCode', 'Sheet1').upper()
            sheet.title = course_code

            # Styles
            times_new_roman = Font(name='Times New Roman', size=12)
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Apply default font
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = times_new_roman

            # Column widths
            column_widths = {
                'A': 3.86, 'B': 15.86, 'C': 42.29, 'D': 10.29,
                'E': 8.14, 'F': 8.14, 'G': 10.43, 'H': 22.71
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Merge cells
            merged_cells = ['A2:C2', 'A3:C3', 'A4:C4', 'D2:E2', 'D3:E3', 'D4:E4']
            for merged_cell in merged_cells:
                sheet.merge_cells(merged_cell)
                top_left = sheet[merged_cell.split(":")[0]]  # only style first cell
                top_left.font = bold_font
                top_left.border = thin_border

            # Header values
            sheet['A2'] = f'Programme: {header.get("programme", "").upper()}'
            sheet['A3'] = f'Course Name: {header.get("courseName", "").upper()}'
            sheet['A4'] = f'Course Code: {header.get("courseCode", "").upper()}'
            sheet['D2'] = "Academic Year :"
            sheet['D3'] = "Semester :"
            sheet['D4'] = "Exam Name :"
            sheet['F2'] = header.get('academicYear', '')
            sheet['F3'] = header.get('semester', '')
            sheet['F4'] = header.get('examName', '')

            # Apply border to header area
            for row in sheet['A2:F4']:
                for cell in row:
                    cell.border = thin_border

            # Blank row
            sheet.append([])

            # Table header
            table_headers = ["SNO", "Reg No", "Name of The Students", "Exam Name", "Mark", "Total", "Percentage"]
            sheet.append(table_headers)
            header_row = sheet.max_row
            for col, title in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], table_headers):
                cell = sheet[f'{col}{header_row}']
                cell.font = bold_font
                cell.border = thin_border
                if col != 'C':
                    cell.alignment = center_alignment

            # Student data
            for idx, student in enumerate(students, start=1):
                sheet.append([
                    idx,
                    student.get('regNo', ''),
                    student.get('studentName', '').upper(),
                    header.get('examName', ""),
                    student.get('marks', 0),
                    student.get('total', 0),
                    student.get('percentage', 0)
                ])

            # Format table
            end_row = sheet.max_row
            for row in sheet[f'A{header_row}:G{end_row}']:
                for cell in row:
                    cell.border = thin_border
            for row in sheet[f'C{header_row}:C{end_row}']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')
            for col in ['A', 'B', 'D', 'E', 'F', 'G']:
                for row in sheet[f'{col}{header_row}:{col}{end_row}']:
                    for cell in row:
                        cell.alignment = center_alignment

            stream = io.BytesIO()
            wb.save(stream)
            file_bytes = stream.getvalue()
            filename = f"{course_code}.xlsx"

            CollegeExamExcel.objects.update_or_create(
                user=request.user,
                course_code=course_code,
                defaults={
                    "file_name": filename,
                    "file_data": file_bytes
                }
            )

            return JsonResponse({"status": "success", "message": "Data stored successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    return render(request, 'staff/college_exam.html')

@login_required
def delete_user(request, user_id):
    dept = get_object_or_404(Department, id=user_id)
    # Find the staff profile linked to this department
    staff_profile = StaffProfile.objects.filter(department=dept).first()
    if staff_profile:
        # Delete the staff user first (this also deletes StaffProfile automatically because of CASCADE)
        staff_user = staff_profile.user
        staff_user.delete()
    # Now delete the department itself
    dept.delete()
    messages.success(request, f"Department '{dept.name}' and its staff profile have been deleted successfully.")
    return redirect('add_dept')



@login_required
def send_merge_file(request):
    if request.method == "POST":
        file_id = request.POST.get("id")
        category_name = request.POST.get("category_name")
        model_name = request.POST.get("model_name")
        dept = request.POST.get("dept")
        if category_name == "COLLEGE EXAM":
            messages.error(request, "College file not able to merge")
            return redirect("department")
        if category_name == "INTERNAL":
            messages.error(request, "Internal file not able to merge")
            return redirect("department")

        source_obj = None
        model_map = {
            'MergeFile': MergeFile,
            'CollegeExamExcel': CollegeExamExcel,
            'InternalExcelFile': InternalExcelFile,
            'ExcelFile': ExcelFile,
        }
        model = model_map.get(model_name)

        try:
            source_obj = model.objects.get(id=file_id, user=request.user)
        except InternalExcelFile.DoesNotExist:
            messages.error(request, "File not found in source tables")
            return redirect("department")
        user = request.user
        if category_name == "THEORY":
            SendMergeFile.objects.update_or_create(
                user=user,
                department=dept,
                file_name=source_obj.file_name,
                defaults={
                    "file_data": source_obj.file_data,
                }
            )
        messages.success(request, "File Added successfully!")
        return redirect("department")
    return redirect("department")

@login_required
def staff_view_merge_file(request):
    staff_profile = request.user.staffprofile
    department_name = staff_profile.department.name
    # Get search and file category from query parameters
    search = request.GET.get('search', "").strip()
    data = SendMergeFile.objects.filter(user=request.user)
    if search:
        data = data.filter(file_name__icontains=search)
    # Order by latest first
    data = data.order_by("-id")
    # Render the page
    return render(request, 'staff/view_merge_excel_file.html', {"dept": department_name,"data": data,"value": search})

def delete_file(request, model_name, file_id):
    if not request.user.is_authenticated:
        messages.error(request, "You must be logged in to delete files.")
        return redirect('frontpage')

    model_map = {
        'MergeFile': MergeFile,
        'SendMergeFile': SendMergeFile,
        'CollegeExamExcel': CollegeExamExcel,
        'InternalExcelFile': InternalExcelFile,
        'ExcelFile': ExcelFile,
    }

    model = model_map.get(model_name)
    if not model:
        messages.error(request, "Invalid file type.")
        return redirect(request.META.get('HTTP_REFERER', '/'))
    try:
        file_obj = model.objects.get(id=file_id, user=request.user)
        file_name=file_obj.file_name
        file_obj.delete()
        messages.success(request, f"File '{file_name}' successfully deleted.")
    except model.DoesNotExist:
        messages.error(request, "File not found or already deleted.")

    # Redirect back to the page that sent the request
    return redirect(request.META.get('HTTP_REFERER', '/'))

def delete_file_from_admin(request, file_id, category):
    if not request.user.is_authenticated:
        messages.error(request, "You must be logged in to delete files.")
        return redirect('frontpage')
    username_prefix = f"{request.user.username}_"
    if category=="SendMergeFile":
        try:
            file_obj = SendMergeFile.objects.get(id=file_id,user__username__startswith=username_prefix)
            file_name = file_obj.file_name
            print(f"File name is : {file_name}")
            file_obj.delete()
            messages.success(request, f"File '{file_name}' successfully deleted.")
        except SendMergeFile.DoesNotExist:
            messages.error(request, "File not found or already deleted.")
    return redirect(request.META.get('HTTP_REFERER', '/'))



@login_required
def admin_merge_file(request):
    username_prefix = f"{request.user.username}_"
    # Get search & category from GET params (better for filtering)
    search = request.GET.get("search", "").strip()
    data = SendMergeFile.objects.filter(user__username__startswith=username_prefix)
    # Apply search filter if any
    if search:
        data = data.filter(file_name__icontains=search)
    # Order result
    data = data.order_by("department", "-id")
    # Render page
    return render(
        request,
        "staff/admin_merge_file.html",
        context={"data": data, "value": search}
    )

def merge_send_files(request):
    user = request.user
    if not user.is_authenticated:
        return redirect("login")

    username_prefix = f"{request.user.username}_"
    send_files = SendMergeFile.objects.filter(user__username__startswith=username_prefix).order_by("-created_at")
    if not send_files.exists():
        messages.error(request, "No files found to merge.")
        return redirect('add_dept')

    try:
        merged_workbook = Workbook()
        if "Sheet" in merged_workbook.sheetnames:
            merged_workbook.remove(merged_workbook["Sheet"])

        # Merge each file
        for send_file in send_files:
            file_stream = io.BytesIO(send_file.file_data)
            source_workbook = load_workbook(file_stream)
            source_sheet = source_workbook.active

            # Collect student data for formatting
            student_data = []
            for row in source_sheet.iter_rows(min_row=7, values_only=True):
                if row and row[0]:
                    student_data.append({
                        'regNo': row[0],
                        'name': row[1],
                        'ese': row[2],
                        'cia': row[3],
                        'total': row[4],
                    })

            # Sheet name
            sheet_name = send_file.file_name[:31]
            new_sheet = merged_workbook.create_sheet(title=sheet_name)

            # Copy data
            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            # Apply formatting
            format_sheet(new_sheet, student_data)

        # Save merged file in memory
        current_time = datetime.now().strftime("%Y-%m-%d")
        output_filename = f"{request.user} merged_{current_time}.xlsx"
        file_stream = io.BytesIO()
        merged_workbook.save(file_stream)
        file_stream.seek(0)

        MergeFile.objects.create(
            user=user,
            file_name=output_filename,
            file_data=file_stream.getvalue()
        )

        messages.success(request, f"File successfully merged and saved.")
        return redirect('add_dept')

    except Exception as e:
        messages.error(request, f"Error while merging files: {str(e)}")
    return redirect('admin_merge_file')


def format_sheet(sheet, student_data=[]):
    # Define styles
    times_new_roman_font = Font(name="Times New Roman", size=12)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply the font to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_new_roman_font

    # Set column widths
    column_widths = {
        "A": 3.86, "B": 15.86, "C": 42.29, "D": 10.29, "E": 8.14,
        "F": 8.14, "G": 9.43, "H": 24.71, "I": 14, "J": 9.43,
        "K": 9.43, "L": 9.43, "M": 9.43, "N": 15.29,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    sheet.merge_cells('A2:C2')
    sheet.merge_cells('A3:C3')
    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D2:E2')
    sheet.merge_cells('D3:E3')
    sheet.merge_cells('D4:E4')
    sheet.merge_cells('H6:I6')
    sheet.merge_cells('H23:K23')
    sheet.merge_cells('H28:K29')
    sheet.merge_cells('L28:L30')
    sheet.merge_cells('M28:M30')
    sheet.merge_cells('N28:N30')

    merge_ranges = ['H50:J50', 'H51:J51', 'H52:J52', 'H53:J53', 'H57:K57']
    for merge_range in merge_ranges:
        sheet.merge_cells(merge_range)

        # Example header formatting (adapt based on your sheet content)
    sheet['A6'].font = bold_font
    sheet['B6'].font = bold_font
    sheet['C6'].font = bold_font
    sheet['D6'].font = bold_font
    sheet['E6'].font = bold_font
    sheet['F6'].font = bold_font

    # Apply center alignment for headers
    for col in ['A', 'B', 'D', 'E', 'F']:
        cell = sheet[f'{col}6']
        cell.alignment = center_alignment

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply borders to merged cells
    merged_cells = ['A2:C2', 'A3:C3', 'A4:C4', 'D2:E2', 'D3:E3', 'D4:E4']
    for merged_cell in merged_cells:
        for row in sheet[merged_cell]:
            for cell in row:
                cell.border = thin_border

    # Apply borders to header section
    for row in sheet['A2:F4']:
        for cell in row:
            cell.border = thin_border

    # Apply borders to student data
    end_row = len(student_data) + 6  # Dynamically calculate the range
    for row in sheet[f'A6:F{end_row}']:
        for cell in row:
            cell.border = thin_border

    # Apply borders to ESE section
    for row in sheet['H6:I21']:
        for cell in row:
            cell.border = thin_border

    # Apply borders to CIA section
    for row in sheet['H23:K25']:
        for cell in row:
            cell.border = thin_border

    # Apply borders to CO Survey section
    for row in sheet['H28:N36']:
        for cell in row:
            cell.border = thin_border

    # Apply borders to Overall Attainment section
    for row in sheet['H39:N48']:
        for cell in row:
            cell.border = thin_border

def merged_file(request):
    if not request.user.is_authenticated:
        return redirect("login")
    data=MergeFile.objects.filter(user=request.user).order_by('-created_at')
    if request.method == 'POST':
        search = request.POST.get('search', "").strip()
        qs = MergeFile.objects.filter(user=request.user)
        if search:
            qs = qs.filter(file_name__icontains=search).order_by('-created_at')
        data = qs
    else:
        search=""
    return render(request, 'staff/admin_merged_files.html',context={"data":data,'value':search})

def custom_logout(request):
    logout(request)
    return redirect("frontpage")

@login_required
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Your password has been changed successfully.")
        else:
            # Collect detailed error messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f" {error}")
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def failed_students(request):
    selected_dept = request.GET.get("departments", "").strip().replace(" ", "_").lower()
    selected = request.GET.get("departments", "").strip()
    status = request.GET.get("status", "Fail")
    query = request.GET.get("query", "").strip()

    username_prefix = f"{request.user.username}_{selected_dept}"
    dept = Department.objects.filter(user=request.user).order_by("name")


    if status == 'All':
        students = StudentGrade.objects.filter(user__username=username_prefix).order_by("reg_no")
        college = CollegeExam.objects.filter(user__username=username_prefix).order_by("reg_no")

    else:
        students = StudentGrade.objects.filter(user__username=username_prefix, status=status).order_by("reg_no")
        college = CollegeExam.objects.filter(user__username=username_prefix, status=status).order_by("reg_no")

    if query:
        students = students.filter(reg_no=query)
        college = college.filter(reg_no=query)

    return render(
        request,
        "staff/failed_student.html",
        {
            "query": query,
            "students": students,
            "college": college,
            "selected_dept": selected,
            "status": status,
            "dept": dept,
        },
    )

@login_required
def dept_failed_students(request):
    status = request.GET.get("status", "Fail")
    query = request.GET.get("query", "").strip()
    if status == 'All':
        students = StudentGrade.objects.filter(user=request.user).order_by("reg_no")
        college = CollegeExam.objects.filter(user=request.user).order_by("reg_no")
    else:
        students = StudentGrade.objects.filter(user=request.user, status=status).order_by("reg_no")
        college = CollegeExam.objects.filter(user=request.user, status=status).order_by("reg_no")
    if query:
        students = students.filter(reg_no=query)
        college = college.filter(reg_no=query)
    return render(request, "staff/dept_failed_students.html", {"query":query,"students": students,"college": college,"status":status})


def student_result(request):
    # query = request.GET.get("query", "").strip()
    # status = request.GET.get("status", "All")
    # students = None
    # college = None
    # if query:
    #     if status != "All":
    #         students = StudentGrade.objects.filter(reg_no=query, status=status).order_by("-total")
    #         college = CollegeExam.objects.filter(reg_no=query, status=status).order_by("-total")
    #     else:
    #         students = StudentGrade.objects.filter(reg_no=query).order_by("-total")
    #         college = CollegeExam.objects.filter(reg_no=query).order_by("-total")
    # context = {"query": query,"status": status,"students": students,"college": college}
    return render(request, "staff/student_result.html")


